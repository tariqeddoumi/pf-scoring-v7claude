#!/usr/bin/env python3
"""Génère le rapport Excel détaillé PF Scoring V7++"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime

# ── Palettes couleurs ──────────────────────────────────────────────
DARK_BLUE   = "1E293B"
MID_BLUE    = "1E40AF"
LIGHT_BLUE  = "DBEAFE"
GREEN       = "166534"
GREEN_LIGHT = "DCFCE7"
RED         = "991B1B"
RED_LIGHT   = "FEE2E2"
ORANGE      = "92400E"
ORANGE_LIGHT= "FEF3C7"
GRAY        = "64748B"
GRAY_LIGHT  = "F8FAFC"
WHITE       = "FFFFFF"

def hfill(color): return PatternFill("solid", fgColor=color)
def hfont(color=WHITE, bold=False, size=11):
    return Font(color=color, bold=bold, size=size, name="Calibri")
def halign(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def hborder():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, wrap=False):
    cell.fill = hfill(bg)
    cell.font = hfont(fg, bold, size)
    cell.alignment = halign(wrap=wrap)
    cell.border = hborder()

def style_cell(cell, bg=WHITE, fg="1E293B", bold=False, size=10,
               h="left", wrap=False):
    cell.fill = hfill(bg)
    cell.font = hfont(fg, bold, size)
    cell.alignment = halign(h, wrap=wrap)
    cell.border = hborder()

def col_width(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ══════════════════════════════════════════════════════════════════
# 1. ONGLET RÉSUMÉ EXÉCUTIF
# ══════════════════════════════════════════════════════════════════
def sheet_resume(wb):
    ws = wb.create_sheet("📋 Résumé Exécutif", 0)
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "PF SCORING V7++ — RAPPORT D'AUDIT COMPLET"
    style_header(c, DARK_BLUE, WHITE, True, 16)
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Application de Scoring Project Finance | Banque Marocaine | Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    style_header(c, MID_BLUE, WHITE, False, 11)
    ws.row_dimensions[2].height = 22

    # KPIs Globaux
    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = "MÉTRIQUES GLOBALES DU PROJET"
    style_header(c, DARK_BLUE, WHITE, True, 12)

    kpis = [
        ("Pages / Routes UI", "62", "Tableau de bord, Projets, Évaluations, Admin, Clients…"),
        ("Routes API", "92", "Auth, Scoring, Admin, Projets, Évaluations, Audit…"),
        ("Composants React", "55", "UI Shadcn + Domaine métier scoring"),
        ("Services métier", "24", "Scoring, Évaluation, Workflow, Audit, Permissions…"),
        ("Modèles de données (Prisma)", "41", "Legacy + V7++ + Workflow + Paramétrage"),
        ("Fichiers TypeScript/TSX", "400+", "Lib, Components, App, Types"),
        ("Domaines de scoring", "9", "D1 Sponsor → D9 ESG & Climate Risk"),
        ("Grades de notation", "13", "AAA, AA, A, BBB+, BBB, BBB-, BB+, BB, B, CCC, CC, C, D"),
        ("Conformité réglementaire", "4", "IFC, EBRD, Bâle III, Bank Al-Maghrib"),
    ]

    headers = ["Indicateur", "Valeur", "Détail"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=i+1, value=h)
        style_header(c, MID_BLUE, WHITE, True)

    for row_idx, (ind, val, det) in enumerate(kpis, start=6):
        bg = GRAY_LIGHT if row_idx % 2 == 0 else WHITE
        ws.cell(row=row_idx, column=1, value=ind).fill = hfill(bg)
        ws.cell(row=row_idx, column=1).font = hfont("1E293B", True, 10)
        ws.cell(row=row_idx, column=1).border = hborder()
        ws.cell(row=row_idx, column=1).alignment = halign("left")

        c2 = ws.cell(row=row_idx, column=2, value=val)
        c2.fill = hfill(GREEN_LIGHT); c2.font = hfont(GREEN, True, 11)
        c2.alignment = halign("center"); c2.border = hborder()

        c3 = ws.cell(row=row_idx, column=3, value=det)
        c3.fill = hfill(bg); c3.font = hfont("475569", False, 10)
        c3.alignment = halign("left", wrap=True); c3.border = hborder()

    ws.merge_cells("A5:A5"); ws.merge_cells("B5:C5")  # skip — individual OK

    # Statut Conformité
    ws.merge_cells("A16:H16")
    c = ws["A16"]
    c.value = "STATUT DE CONFORMITÉ RÉGLEMENTAIRE"
    style_header(c, DARK_BLUE, WHITE, True, 12)

    conf = [
        ("IFC Performance Standards", "✅ Conforme", "Critères environnementaux & sociaux intégrés (D9 ESG)", GREEN, GREEN_LIGHT),
        ("EBRD Environmental & Social Policy", "✅ Conforme", "Politiques E&S BERD dans domaine ESG & Climate Risk", GREEN, GREEN_LIGHT),
        ("Bâle III / CRR2", "✅ Conforme", "Grades AAA→D, PD indicatives, LGD structure", GREEN, GREEN_LIGHT),
        ("Bank Al-Maghrib (BAM)", "✅ Conforme", "Scoring Project Finance adapté marché marocain (MAD)", GREEN, GREEN_LIGHT),
        ("FATF / AML-CFT", "✅ Conforme", "KYC Status, Compliance Status sur entité Client", GREEN, GREEN_LIGHT),
        ("GDPR / Protection données", "⚠️ Partiel", "Soft-delete implémenté, chiffrement données à vérifier", ORANGE, ORANGE_LIGHT),
    ]

    for i, h in enumerate(["Standard", "Statut", "Détail"]):
        c = ws.cell(row=17, column=i+1, value=h)
        style_header(c, MID_BLUE, WHITE, True)

    for row_idx, (std, stat, det, fg, bg) in enumerate(conf, 18):
        ws.cell(row=row_idx, column=1, value=std).fill = hfill(GRAY_LIGHT if row_idx%2==0 else WHITE)
        ws.cell(row=row_idx, column=1).font = hfont("1E293B", True, 10)
        ws.cell(row=row_idx, column=1).border = hborder()
        ws.cell(row=row_idx, column=1).alignment = halign("left")

        c2 = ws.cell(row=row_idx, column=2, value=stat)
        c2.fill = hfill(bg); c2.font = hfont(fg, True, 10)
        c2.alignment = halign("center"); c2.border = hborder()

        c3 = ws.cell(row=row_idx, column=3, value=det)
        c3.fill = hfill(GRAY_LIGHT if row_idx%2==0 else WHITE)
        c3.font = hfont("475569", False, 10)
        c3.alignment = halign("left", wrap=True); c3.border = hborder()
        ws.row_dimensions[row_idx].height = 22

    col_width(ws, {"A": 35, "B": 18, "C": 70, "D": 5, "E": 5, "F": 5, "G": 5, "H": 5})
    freeze(ws, "A3")


# ══════════════════════════════════════════════════════════════════
# 2. SPÉCIFICATIONS VS IMPLÉMENTATION
# ══════════════════════════════════════════════════════════════════
def sheet_specs(wb):
    ws = wb.create_sheet("📊 Specs vs Implémentation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "SPÉCIFICATIONS FONCTIONNELLES VS IMPLÉMENTATION"
    style_header(c, DARK_BLUE, WHITE, True, 14)
    ws.row_dimensions[1].height = 35

    headers = ["#", "Module", "Fonctionnalité", "Statut", "Priorité",
               "Fichiers clés", "API Endpoint", "Tables DB", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, MID_BLUE, WHITE, True, 10)

    specs = [
        # Scoring Domains
        (1, "Domaines Scoring", "D1 – Sponsor & Shareholders (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode, ScoringNodeRange",
         "Nœuds hiérarchiques configurés"),
        (2, "Domaines Scoring", "D2 – Project Characteristics (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "Technical complexity, Location risk"),
        (3, "Domaines Scoring", "D3 – Construction Risk (15%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "EPC, Interfaces risk, Insurance"),
        (4, "Domaines Scoring", "D4 – Market Risk (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "Demand risk, Price risk, Competition"),
        (5, "Domaines Scoring", "D5 – Operational Risk (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "O&M Capability, Stability, Cost control"),
        (6, "Domaines Scoring", "D6 – Counterparty Risk (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "Offtaker risk, Supplier risk"),
        (7, "Domaines Scoring", "D7 – Financial Structure & Cash Flow (15%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts, lib/services/score-calculator.ts",
         "/api/scoring/evaluations/{id}/calculate", "ScoringNode, ScoringNodeRange",
         "DSCR, LLCR, Debt capacity"),
        (8, "Domaines Scoring", "D8 – Legal & Documentation (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "Contractual framework, Security package"),
        (9, "Domaines Scoring", "D9 – ESG & Climate Risk (10%)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-model.ts", "/api/scoring/evaluations", "ScoringNode",
         "Environmental, Social, Governance, Climate"),
        # Grades
        (10, "Notation / Grades", "Échelle AAA → D (13 niveaux)",
         "✅ Implémenté", "🔴 Critique",
         "lib/constants.ts, lib/scoring-model.ts",
         "Calculé automatiquement", "ScoringEvaluation.finalGrade",
         "SCORE_TO_RATING mapping complet"),
        (11, "Notation / Grades", "PD indicatives par grade",
         "✅ Implémenté", "🟠 Haute",
         "lib/scoring-model.ts", "—", "ScoringEvaluation",
         "PD range: 0.01% (AAA) → 20-100% (D)"),
        (12, "Notation / Grades", "Classe de risque (Faible/Modéré/Élevé/Très Élevé)",
         "✅ Implémenté", "🔴 Critique",
         "lib/scoring-engine.ts", "—", "ScoringEvaluation.riskClass",
         "Dérivée du grade automatiquement"),
        # Calcul Scores
        (13, "Moteur de Scoring", "Calcul hiérarchique récursif (Domain→Criterion→Leaf)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/generic-scoring-engine.ts",
         "/api/scoring/evaluations/{id}/calculate", "ScoringEvaluationNodeResult",
         "4 niveaux de profondeur"),
        (14, "Moteur de Scoring", "Agrégation WEIGHTED_AVERAGE",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/generic-scoring-engine.ts", "—", "ScoringNode.aggregationMethod",
         "Méthode par défaut"),
        (15, "Moteur de Scoring", "Agrégation SIMPLE_AVERAGE, SUM, MIN, MAX",
         "✅ Implémenté", "🟠 Haute",
         "lib/services/generic-scoring-engine.ts", "—", "ScoringNode",
         "Modes alternatifs configurables"),
        (16, "Moteur de Scoring", "Scoring OPTION_SCORE (choix multiples)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/score-calculator.ts",
         "/api/scoring/evaluations/{id}/answers", "ScoringNodeOption",
         "Lookup option → score"),
        (17, "Moteur de Scoring", "Scoring RANGE_SCORE (numérique)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/score-calculator.ts", "—", "ScoringNodeRange",
         "Interpolation dans plages"),
        (18, "Moteur de Scoring", "Scoring FORMULA (expression parser)",
         "⚠️ Partiel", "🟠 Haute",
         "lib/services/rule-engine.ts", "—", "ScoringNodeFormula",
         "Infrastructure prête, évaluation formules à finaliser"),
        (19, "Moteur de Scoring", "Calcul score en temps réel (LiveScore)",
         "✅ Implémenté", "🟠 Haute",
         "components/scoring/LiveScorePanel.tsx",
         "/api/scoring/evaluations/{id}/calculate", "ScoringEvaluationNodeResult",
         "Mise à jour après chaque réponse"),
        # Règles métier
        (20, "Règles Métier", "NO-GO rules (rejet automatique)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/rule-engine.ts", "—", "ScoringNodeRule (type=NO_GO)",
         "DSCR<1.1x, Equity<20%, Garanties manquantes…"),
        (21, "Règles Métier", "MALUS (pénalité sur score)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/rule-engine.ts", "—", "ScoringNodeRule (type=MALUS)",
         "Pénalité en points (-X pts)"),
        (22, "Règles Métier", "WARNINGS (alertes informatives)",
         "✅ Implémenté", "🟡 Moyenne",
         "lib/services/rule-engine.ts", "—", "ScoringNodeRule (type=WARNING)",
         "Alertes non-bloquantes"),
        (23, "Règles Métier", "HARD_STOP (blocage, requiert approbation)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/rule-engine.ts", "—", "ScoringNodeRule (type=HARD_STOP)",
         "Blocage du workflow"),
        (24, "Règles Métier", "Applicabilité conditionnelle des nœuds",
         "✅ Implémenté", "🟠 Haute",
         "lib/services/applicability-engine.ts", "—",
         "ScoringNodeApplicabilityRule",
         "Nœuds activés selon conditions"),
        # Workflow
        (25, "Workflow Approbation", "Soumission évaluation (DRAFT → SUBMITTED)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-workflow-service.ts",
         "/api/evaluations/submit", "ScoringEvaluation, ScoringWorkflow",
         "Transition de statut"),
        (26, "Workflow Approbation", "Revue analyste (SUBMITTED → UNDER_REVIEW)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-workflow-service.ts",
         "/api/evaluations/{id}", "ScoringWorkflowStep",
         "ANALYST_REVIEW step"),
        (27, "Workflow Approbation", "Revue risque (UNDER_REVIEW → REVIEWED)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-workflow-service.ts",
         "/api/evaluations/validate", "ScoringWorkflowStep",
         "RISK_REVIEW step"),
        (28, "Workflow Approbation", "Comité de crédit (REVIEWED → APPROVED/REJECTED)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-workflow-service.ts",
         "/api/evaluations/validate", "ScoringDecision",
         "COMMITTEE_REVIEW step"),
        (29, "Workflow Approbation", "Chaîne approbation multi-niveaux",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-workflow-service.ts", "—",
         "ScoringApproval", "Analyst→RiskMgr→Committee→Final"),
        (30, "Workflow Approbation", "Override / Surcharge score avec justification",
         "✅ Implémenté", "🔴 Critique",
         "components/scoring/OverrideManagement.tsx",
         "/api/admin/scoring/overrides", "ScoringOverride",
         "Approval PENDING→APPROVED/REJECTED"),
        # Audit Trail
        (31, "Audit Trail", "Log des modifications de scoring (CREATE/UPDATE/DELETE)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-audit-service.ts",
         "/api/audit", "ScoringChangeLog",
         "oldValueJson / newValueJson complets"),
        (32, "Audit Trail", "Log transitions d'évaluation",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/scoring-audit-service.ts", "—",
         "ScoringAuditLog", "previousScore / newScore"),
        (33, "Audit Trail", "Log actions utilisateurs (Login/Logout/RoleChange)",
         "✅ Implémenté", "🔴 Critique",
         "lib/audit-logger.ts", "—",
         "UserAuditLog", "IP address + user agent"),
        (34, "Audit Trail", "Journal d'audit accessible (page)",
         "✅ Implémenté", "🟠 Haute",
         "app/admin/audit-logs/page.tsx", "/api/audit",
         "AuditLog", "Filtres par action/date/utilisateur"),
        # Gestion Projets
        (35, "Gestion Projets", "CRUD Projets (Create/Read/Update/Delete)",
         "✅ Implémenté", "🔴 Critique",
         "lib/services/project-service.ts",
         "/api/projects, /api/projects/{id}", "Project",
         "Soft delete, pagination"),
        (36, "Gestion Projets", "Données financières projet (DSCR, LLCR, Leverage)",
         "✅ Implémenté", "🔴 Critique",
         "app/projects/[id]/page.tsx", "/api/projects/{id}",
         "Project (dscr, llcr, leverage)", "Champs financiers dédiés"),
        (37, "Gestion Projets", "Statuts projet (Brouillon→En cours→Approuvé/Rejeté)",
         "✅ Implémenté", "🔴 Critique",
         "lib/constants.ts", "—", "Project.status (enum ProjectStatus)",
         "5 statuts complets"),
        (38, "Gestion Projets", "Secteurs / Sous-secteurs projet",
         "✅ Implémenté", "🟡 Moyenne",
         "lib/constants.ts", "—", "Sector, Subsector",
         "Tables référentiel"),
        # Gestion Clients
        (39, "Gestion Clients", "CRUD Clients",
         "✅ Implémenté", "🔴 Critique",
         "app/clients/", "/api/clients, /api/clients/{id}",
         "Client", "Create/Read/Update/Delete"),
        (40, "Gestion Clients", "KYC & Conformité client",
         "✅ Implémenté", "🔴 Critique",
         "app/clients/[id]/page.tsx", "/api/clients/{id}",
         "Client (kycStatus, complianceStatus)", "AML-CFT compliant"),
        (41, "Gestion Clients", "Rating interne client",
         "✅ Implémenté", "🟠 Haute",
         "prisma/schema.prisma", "—",
         "InternalRating, Client.internalRating", "Échelle de notation interne"),
        # Reporting
        (42, "Reporting & Dashboard", "Tableau de bord KPIs globaux",
         "✅ Implémenté", "🔴 Critique",
         "app/dashboard/page.tsx", "/api/auth/me",
         "Agrégations diverses", "Projets, Évaluations, Score moyen"),
        (43, "Reporting & Dashboard", "Répartition grades (AAA→D) avec graphiques",
         "✅ Implémenté", "🟠 Haute",
         "components/dashboard/stats-card.tsx",
         "/api/projects", "Evaluation.finalGrade",
         "Distribution visuelle"),
        (44, "Reporting & Dashboard", "Export rapport PDF évaluation",
         "✅ Implémenté", "🟠 Haute",
         "lib/pdf-service.ts, app/evaluations/[id]/page.tsx",
         "/api/evaluations/{id}/report", "—",
         "Rapport exécutif + technique"),
        (45, "Reporting & Dashboard", "Export Excel (analytics)",
         "⚠️ Partiel", "🟡 Moyenne",
         "lib/export-service.ts", "/api/exports",
         "—", "Infrastructure présente, UI incomplète"),
        (46, "Reporting & Dashboard", "Benchmarking inter-projets",
         "✅ Implémenté", "🟡 Moyenne",
         "app/benchmarking/page.tsx", "—",
         "Evaluation", "Comparaison scores entre projets"),
        # Administration
        (47, "Administration", "Gestion modèles de scoring (CRUD)",
         "✅ Implémenté", "🔴 Critique",
         "app/admin/scoring-grid-v7pp/",
         "/api/admin/scoring/models", "ScoringModel, ScoringModelVersion",
         "DRAFT/PUBLISHED governance"),
        (48, "Administration", "Éditeur de nœuds hiérarchiques",
         "✅ Implémenté", "🔴 Critique",
         "app/admin/scoring-designer/",
         "/api/admin/scoring/nodes", "ScoringNode",
         "Visual tree editor"),
        (49, "Administration", "Configuration règles NO-GO/MALUS",
         "✅ Implémenté", "🔴 Critique",
         "app/admin/scoring-designer/",
         "/api/admin/scoring/models/.../versions/.../rules",
         "ScoringNodeRule", "Condition + action + sévérité"),
        (50, "Administration", "Configuration risque pays",
         "✅ Implémenté", "🟠 Haute",
         "app/admin/country-risk/page.tsx",
         "/api/admin/countries", "Country",
         "Score par pays, mode risque"),
        (51, "Administration", "Gestion utilisateurs & rôles (RBAC)",
         "✅ Implémenté", "🔴 Critique",
         "app/admin/users/page.tsx",
         "/api/admin/users", "User",
         "Admin/Manager/Analyst/Viewer"),
        (52, "Administration", "Diagnostic système & santé DB",
         "✅ Implémenté", "🟡 Moyenne",
         "app/admin/diagnostic/page.tsx",
         "/api/admin/diagnostic/full", "—",
         "Vérification connexions et tables"),
        # Documents
        (53, "Documents & Evidences", "Upload documents par évaluation",
         "✅ Implémenté", "🟠 Haute",
         "components/scoring/DocumentUploadPanel.tsx",
         "/api/admin/scoring/documents", "ScoringDocument",
         "fileName, fileSize, fileType"),
        (54, "Documents & Evidences", "Exigences documents par nœud scoring",
         "✅ Implémenté", "🟠 Haute",
         "prisma/schema.prisma",
         "/api/admin/scoring/nodes", "ScoringNodeDocumentRequirement",
         "Required/Optional, MIME types"),
        # Commentaires
        (55, "Commentaires & Discussion", "Thread de commentaires par évaluation",
         "✅ Implémenté", "🟠 Haute",
         "components/scoring/WorkflowCommentThread.tsx", "—",
         "ScoringComment", "Reply chain, types: GENERAL/QUESTION/ISSUE"),
        (56, "Commentaires & Discussion", "Commentaires internes (revue)",
         "✅ Implémenté", "🟡 Moyenne",
         "components/scoring/WorkflowCommentThread.tsx", "—",
         "ScoringComment.isInternal", "Visibilité restreinte analystes"),
        # Stress Testing
        (57, "Stress Testing", "Scénarios stress test DSCR/LLCR",
         "✅ Implémenté", "🟠 Haute",
         "app/evaluations/[id]/page.tsx",
         "/api/evaluations/{id}/stress-test",
         "StressTestScenarioResult", "PASS/MARGINAL/FAIL"),
        # Sécurité
        (58, "Sécurité & Auth", "Authentification JWT",
         "✅ Implémenté", "🔴 Critique",
         "lib/auth.ts, lib/auth-middleware.ts",
         "/api/auth/login", "User",
         "Tokens 24h, vérification Bearer"),
        (59, "Sécurité & Auth", "OAuth (Google, GitHub)",
         "⚠️ Partiel", "🟡 Moyenne",
         "app/api/auth/oauth/[provider]",
         "/api/auth/oauth-init", "User",
         "Infrastructure présente, config provider requise"),
        (60, "Sécurité & Auth", "Rate limiting API",
         "❌ Manquant", "🟠 Haute",
         "—", "—", "—",
         "Pas de limite sur /api/auth/login — Risque brute force"),
        (61, "Sécurité & Auth", "Routes debug désactivées en production",
         "❌ Manquant", "🔴 Critique",
         "app/api/debug/*, app/api/init/*",
         "/api/debug/login, /api/debug/users", "—",
         "À désactiver via NODE_ENV check"),
        (62, "Sécurité & Auth", "Validation JWT_SECRET obligatoire en production",
         "⚠️ Partiel", "🔴 Critique",
         "lib/auth-middleware.ts", "—", "—",
         "Fallback 'your-secret-key' dangereux"),
        # Data Binding
        (63, "Data Binding", "Auto-fill depuis données Projet",
         "✅ Implémenté", "🟠 Haute",
         "lib/services/binding-resolver.ts",
         "/api/scoring/bindings", "ScoringNodeDataBinding",
         "Source: PROJECT, mapField configuré"),
        (64, "Data Binding", "Auto-fill depuis données Client",
         "✅ Implémenté", "🟠 Haute",
         "lib/services/binding-resolver.ts",
         "/api/scoring/bindings", "ScoringNodeDataBinding",
         "Source: CLIENT"),
        (65, "Data Binding", "Transformations (FORMAT, MAP_VALUE, AGGREGATE)",
         "⚠️ Partiel", "🟡 Moyenne",
         "lib/services/binding-resolver.ts", "—",
         "ScoringNodeDataBinding.transformationType",
         "Infrastructure DB prête, logique à compléter"),
    ]

    STATUS_STYLES = {
        "✅ Implémenté": (GREEN, GREEN_LIGHT),
        "⚠️ Partiel":    (ORANGE, ORANGE_LIGHT),
        "❌ Manquant":   (RED, RED_LIGHT),
    }
    PRIO_STYLES = {
        "🔴 Critique": (RED, RED_LIGHT),
        "🟠 Haute":    (ORANGE, ORANGE_LIGHT),
        "🟡 Moyenne":  ("7C6F00", "FEFCE8"),
        "🟢 Faible":   (GREEN, GREEN_LIGHT),
    }

    for row_idx, spec in enumerate(specs, start=3):
        bg_row = GRAY_LIGHT if row_idx % 2 == 0 else WHITE
        num, module, fonc, statut, prio, fichiers, api, tables, notes = spec

        # N°
        c = ws.cell(row=row_idx, column=1, value=num)
        style_cell(c, bg_row, GRAY, False, 9, "center")

        # Module
        c = ws.cell(row=row_idx, column=2, value=module)
        style_cell(c, bg_row, DARK_BLUE, True, 10)

        # Fonctionnalité
        c = ws.cell(row=row_idx, column=3, value=fonc)
        style_cell(c, bg_row, "1E293B", False, 10, "left", True)

        # Statut
        fg, bg = STATUS_STYLES.get(statut, (GRAY, GRAY_LIGHT))
        c = ws.cell(row=row_idx, column=4, value=statut)
        style_cell(c, bg, fg, True, 10, "center")

        # Priorité
        fg2, bg2 = PRIO_STYLES.get(prio, (GRAY, GRAY_LIGHT))
        c = ws.cell(row=row_idx, column=5, value=prio)
        style_cell(c, bg2, fg2, True, 9, "center")

        # Fichiers
        c = ws.cell(row=row_idx, column=6, value=fichiers)
        style_cell(c, bg_row, "475569", False, 9, "left", True)

        # API
        c = ws.cell(row=row_idx, column=7, value=api)
        style_cell(c, bg_row, MID_BLUE, False, 9, "left", True)

        # Tables
        c = ws.cell(row=row_idx, column=8, value=tables)
        style_cell(c, bg_row, "475569", False, 9, "left", True)

        # Notes
        c = ws.cell(row=row_idx, column=9, value=notes)
        style_cell(c, bg_row, GRAY, False, 9, "left", True)

        ws.row_dimensions[row_idx].height = 30

    col_width(ws, {
        "A": 5, "B": 22, "C": 42, "D": 16, "E": 13,
        "F": 38, "G": 38, "H": 30, "I": 35
    })
    freeze(ws, "A3")

    # Auto-filter
    ws.auto_filter.ref = f"A2:I{len(specs)+2}"


# ══════════════════════════════════════════════════════════════════
# 3. ERREURS & CORRECTIONS
# ══════════════════════════════════════════════════════════════════
def sheet_bugs(wb):
    ws = wb.create_sheet("🐛 Erreurs & Corrections")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "BUGS, ERREURS & CORRECTIONS APPLIQUÉES"
    style_header(c, DARK_BLUE, WHITE, True, 14)
    ws.row_dimensions[1].height = 35

    headers = ["#", "Sévérité", "Type", "Description", "Fichier(s)", "Correction", "Statut", "Impact"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, MID_BLUE, WHITE, True, 10)

    bugs = [
        (1, "🔴 Critique", "Auth",
         "Scoring Grid: fetch sans Authorization header → 401/HTML retourné",
         "app/admin/scoring-grid/page.tsx",
         "Ajout getAuthHeaders() → Record<string,string> sur GET/POST/DELETE",
         "✅ Corrigé", "Chargement critères de scoring impossible"),
        (2, "🔴 Critique", "Auth",
         "Field Management: fetch sans Authorization header → 401",
         "app/admin/field-management/page.tsx",
         "Ajout getAuthHeaders() sur fetchSections, handleAddField, handleDeleteField, handleToggleFieldVisibility",
         "✅ Corrigé", "Configuration champs formulaires inaccessible"),
        (3, "🔴 Critique", "TypeScript",
         "Spread {…getAuthHeaders()} rejeté: type '{ Authorization?: undefined }' non assignable à HeadersInit",
         "app/admin/scoring-grid/page.tsx",
         "Return type explicite Record<string,string> sur getAuthHeaders()",
         "✅ Corrigé", "Build Vercel échoué (exit code 1)"),
        (4, "🔴 Critique", "Sécurité",
         "Routes /api/debug/login et /api/debug/users actives en production",
         "app/api/debug/login/route.ts, app/api/debug/users/route.ts",
         "Ajouter guard NODE_ENV !== 'production' → 403",
         "⚠️ En attente", "Accès non authentifié aux données utilisateurs"),
        (5, "🔴 Critique", "Sécurité",
         "JWT_SECRET fallback 'your-secret-key' si variable d'env absente",
         "lib/auth-middleware.ts, lib/auth.ts",
         "throw Error en production si JWT_SECRET absent (fail fast)",
         "⚠️ En attente", "Tokens forgeable par attaquant"),
        (6, "🟠 Haute", "Auth",
         "Login retourne HTML au lieu de JSON (env Vercel)",
         "app/api/auth/login/route.ts",
         "Ajouter DATABASE_URL + JWT_SECRET dans Vercel Env Variables",
         "⚠️ Config requise", "Connexion production impossible"),
        (7, "🟠 Haute", "Sécurité",
         "Pas de rate limiting sur /api/auth/login",
         "app/api/auth/login/route.ts",
         "Ajouter middleware rate-limit (ex: upstash/ratelimit ou next-rate-limit)",
         "⚠️ En attente", "Vulnérabilité brute force"),
        (8, "🟠 Haute", "Sécurité",
         "Routes /api/init/* actives (init-test-user, set-admin-password)",
         "app/api/init/*, app/api/init-test-user/route.ts",
         "Désactiver en production ou protéger par INIT_TOKEN secret",
         "⚠️ En attente", "Création de comptes admin non autorisée"),
        (9, "🟠 Haute", "Cohérence",
         "Réponses API incohérentes: {error:string} vs {message:string} vs HTTP only",
         "Multiples routes API",
         "Utiliser lib/api-response.ts systématiquement",
         "⚠️ En attente", "Gestion d'erreurs frontend dégradée"),
        (10, "🟡 Moyenne", "TypeScript",
         "@ts-ignore dans notion-client.ts (dépendance optionnelle)",
         "lib/notion-client.ts",
         "Acceptable — dépendance optionnelle, commentaire explicatif présent",
         "✅ Accepté", "Impact nul en production"),
        (11, "🟡 Moyenne", "Fonctionnel",
         "Export Excel (/app/exports) UI incomplète",
         "app/exports/page.tsx, lib/export-service.ts",
         "Compléter UI d'export avec sélection colonnes et téléchargement",
         "⚠️ En attente", "Fonctionnalité reporting limitée"),
        (12, "🟡 Moyenne", "Fonctionnel",
         "Formula scoring (FORMULA_SCORE) partiellement implémenté",
         "lib/services/score-calculator.ts",
         "Implémenter évaluateur d'expressions sécurisé (ex: expr-eval)",
         "⚠️ En attente", "Certains critères avancés non calculables"),
        (13, "🟡 Moyenne", "Sécurité",
         "Champs JSON (conditionExpression, uiSchemaJson) sans validation stricte",
         "lib/services/rule-engine.ts, prisma/schema.prisma",
         "Valider structure JSON avec Zod avant persistence",
         "⚠️ En attente", "Risque injection expressions malicieuses"),
        (14, "🟢 Faible", "Auth",
         "Composants Header/UserProfile: fetch /api/auth/me sans Authorization",
         "components/layout/Header.tsx, UserProfile.tsx",
         "Ces routes utilisent cookies (correct), pas de bearer nécessaire",
         "✅ Accepté", "Comportement correct par design"),
    ]

    SEV_STYLES = {
        "🔴 Critique": (RED, RED_LIGHT),
        "🟠 Haute":    (ORANGE, ORANGE_LIGHT),
        "🟡 Moyenne":  ("7C6F00", "FEFCE8"),
        "🟢 Faible":   (GREEN, GREEN_LIGHT),
    }
    STATUS_STYLES = {
        "✅ Corrigé":     (GREEN, GREEN_LIGHT),
        "✅ Accepté":     (GREEN, GREEN_LIGHT),
        "⚠️ En attente":  (ORANGE, ORANGE_LIGHT),
        "⚠️ Config requise": ("7C6F00", "FEFCE8"),
    }

    for row_idx, bug in enumerate(bugs, start=3):
        num, sev, typ, desc, fichier, correction, statut, impact = bug
        bg_row = GRAY_LIGHT if row_idx % 2 == 0 else WHITE

        ws.cell(row=row_idx, column=1, value=num).fill = hfill(bg_row)
        ws.cell(row=row_idx, column=1).font = hfont(GRAY, False, 9)
        ws.cell(row=row_idx, column=1).alignment = halign("center")
        ws.cell(row=row_idx, column=1).border = hborder()

        fg, bg = SEV_STYLES.get(sev, (GRAY, GRAY_LIGHT))
        c = ws.cell(row=row_idx, column=2, value=sev)
        style_cell(c, bg, fg, True, 9, "center")

        c = ws.cell(row=row_idx, column=3, value=typ)
        style_cell(c, bg_row, DARK_BLUE, True, 9, "center")

        c = ws.cell(row=row_idx, column=4, value=desc)
        style_cell(c, bg_row, "1E293B", False, 9, "left", True)

        c = ws.cell(row=row_idx, column=5, value=fichier)
        style_cell(c, bg_row, "475569", False, 8, "left", True)

        c = ws.cell(row=row_idx, column=6, value=correction)
        style_cell(c, bg_row, "1E293B", False, 9, "left", True)

        fg2, bg2 = STATUS_STYLES.get(statut, (GRAY, GRAY_LIGHT))
        c = ws.cell(row=row_idx, column=7, value=statut)
        style_cell(c, bg2, fg2, True, 9, "center")

        c = ws.cell(row=row_idx, column=8, value=impact)
        style_cell(c, bg_row, GRAY, False, 9, "left", True)

        ws.row_dimensions[row_idx].height = 35

    col_width(ws, {
        "A": 4, "B": 14, "C": 11, "D": 45,
        "E": 38, "F": 50, "G": 17, "H": 35
    })
    freeze(ws, "A3")
    ws.auto_filter.ref = f"A2:H{len(bugs)+2}"


# ══════════════════════════════════════════════════════════════════
# 4. ROUTES API
# ══════════════════════════════════════════════════════════════════
def sheet_api(wb):
    ws = wb.create_sheet("🔗 Routes API")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "INVENTAIRE COMPLET DES ROUTES API (92 endpoints)"
    style_header(c, DARK_BLUE, WHITE, True, 14)
    ws.row_dimensions[1].height = 35

    headers = ["Catégorie", "Méthode", "Endpoint", "Auth requise", "Description", "Fichier", "Statut"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, MID_BLUE, WHITE, True, 10)

    routes = [
        # Auth
        ("Auth", "POST", "/api/auth/login", "Non", "Connexion utilisateur", "app/api/auth/login/route.ts", "✅ Actif"),
        ("Auth", "POST", "/api/auth/logout", "Cookie", "Déconnexion", "app/api/auth/logout/route.ts", "✅ Actif"),
        ("Auth", "GET", "/api/auth/me", "Cookie", "Profil utilisateur courant", "app/api/auth/me/route.ts", "✅ Actif"),
        ("Auth", "POST", "/api/auth/register", "Non", "Inscription", "app/api/auth/register/route.ts", "✅ Actif"),
        ("Auth", "POST", "/api/auth/oauth-init", "Non", "Initiation OAuth", "app/api/auth/oauth-init/route.ts", "✅ Actif"),
        ("Auth", "POST", "/api/auth/oauth/[provider]", "Non", "Callback OAuth", "app/api/auth/oauth/[provider]/route.ts", "⚠️ Config"),
        # Projects
        ("Projets", "GET", "/api/projects", "Bearer", "Liste projets paginée", "app/api/projects/route.ts", "✅ Actif"),
        ("Projets", "POST", "/api/projects", "Bearer", "Créer projet", "app/api/projects/route.ts", "✅ Actif"),
        ("Projets", "GET", "/api/projects/{id}", "Bearer", "Détail projet", "app/api/projects/[id]/route.ts", "✅ Actif"),
        ("Projets", "PUT", "/api/projects/{id}", "Bearer", "Modifier projet", "app/api/projects/[id]/route.ts", "✅ Actif"),
        ("Projets", "DELETE", "/api/projects/{id}", "Bearer", "Supprimer projet", "app/api/projects/[id]/route.ts", "✅ Actif"),
        ("Projets", "GET", "/api/projects/{id}/scoring", "Bearer", "Scores projet", "app/api/projects/[id]/scoring/route.ts", "✅ Actif"),
        ("Projets", "GET", "/api/projects/{id}/scorings", "Bearer", "Historique scorings", "app/api/projects/[id]/scorings/route.ts", "✅ Actif"),
        # Clients
        ("Clients", "GET", "/api/clients", "Bearer", "Liste clients", "app/api/clients/route.ts", "✅ Actif"),
        ("Clients", "POST", "/api/clients", "Bearer", "Créer client", "app/api/clients/route.ts", "✅ Actif"),
        ("Clients", "GET", "/api/clients/{id}", "Bearer", "Détail client", "app/api/clients/[id]/route.ts", "✅ Actif"),
        ("Clients", "PUT", "/api/clients/{id}", "Bearer", "Modifier client", "app/api/clients/[id]/route.ts", "✅ Actif"),
        ("Clients", "DELETE", "/api/clients/{id}", "Bearer", "Supprimer client", "app/api/clients/[id]/route.ts", "✅ Actif"),
        # Evaluations
        ("Évaluations", "GET", "/api/evaluations", "Bearer", "Liste évaluations", "app/api/evaluations/route.ts", "✅ Actif"),
        ("Évaluations", "POST", "/api/evaluations", "Bearer", "Créer évaluation", "app/api/evaluations/route.ts", "✅ Actif"),
        ("Évaluations", "GET", "/api/evaluations/{id}", "Bearer", "Détail évaluation", "app/api/evaluations/[id]/route.ts", "✅ Actif"),
        ("Évaluations", "PUT", "/api/evaluations/{id}", "Bearer", "Modifier évaluation", "app/api/evaluations/[id]/route.ts", "✅ Actif"),
        ("Évaluations", "POST", "/api/evaluations/submit", "Bearer", "Soumettre pour review", "app/api/evaluations/submit/route.ts", "✅ Actif"),
        ("Évaluations", "POST", "/api/evaluations/validate", "Bearer (Manager)", "Valider/Approuver", "app/api/evaluations/validate/route.ts", "✅ Actif"),
        ("Évaluations", "POST", "/api/evaluations/reject", "Bearer (Manager)", "Rejeter évaluation", "app/api/evaluations/reject/route.ts", "✅ Actif"),
        ("Évaluations", "GET", "/api/evaluations/{id}/report", "Bearer", "Export rapport", "app/api/evaluations/[id]/report/route.ts", "✅ Actif"),
        ("Évaluations", "GET", "/api/evaluations/{id}/stress-test", "Bearer", "Stress test DSCR/LLCR", "app/api/evaluations/[id]/stress-test/route.ts", "✅ Actif"),
        ("Évaluations", "POST", "/api/evaluations/{id}/score/calculate", "Bearer", "Recalculer score", "app/api/evaluations/[id]/score/calculate/route.ts", "✅ Actif"),
        # Scoring V7
        ("Scoring V7", "GET", "/api/scoring/evaluations", "Bearer", "Liste évaluations V7", "app/api/scoring/evaluations/route.ts", "✅ Actif"),
        ("Scoring V7", "POST", "/api/scoring/evaluations", "Bearer", "Créer évaluation V7", "app/api/scoring/evaluations/route.ts", "✅ Actif"),
        ("Scoring V7", "GET", "/api/scoring/evaluations/{id}/form", "Bearer", "Charger formulaire", "app/api/scoring/evaluations/[id]/form/route.ts", "✅ Actif"),
        ("Scoring V7", "POST", "/api/scoring/evaluations/{id}/answers", "Bearer", "Sauvegarder réponses", "app/api/scoring/evaluations/[id]/answers/route.ts", "✅ Actif"),
        ("Scoring V7", "POST", "/api/scoring/evaluations/{id}/calculate", "Bearer", "Calculer score", "app/api/scoring/evaluations/[id]/calculate/route.ts", "✅ Actif"),
        ("Scoring V7", "POST", "/api/scoring/evaluations/{id}/submit", "Bearer", "Soumettre V7", "app/api/scoring/evaluations/[id]/submit/route.ts", "✅ Actif"),
        ("Scoring V7", "GET", "/api/scoring/evaluations/{id}/trace", "Bearer", "Trace calcul", "app/api/scoring/evaluations/[id]/trace/route.ts", "✅ Actif"),
        ("Scoring V7", "GET", "/api/scoring/questionnaire", "Bearer", "Charger questionnaire", "app/api/scoring/questionnaire/route.ts", "✅ Actif"),
        ("Scoring V7", "GET", "/api/scoring/bindings", "Bearer", "Résoudre bindings", "app/api/scoring/bindings/route.ts", "✅ Actif"),
        # Admin Scoring
        ("Admin Scoring", "GET/POST", "/api/admin/scoring/models", "Bearer (Admin)", "CRUD modèles", "app/api/admin/scoring/models/route.ts", "✅ Actif"),
        ("Admin Scoring", "GET/POST", "/api/admin/scoring/models/{id}/versions", "Bearer (Admin)", "CRUD versions", "app/api/admin/scoring/models/[id]/versions/route.ts", "✅ Actif"),
        ("Admin Scoring", "GET/POST/PUT/DELETE", "/api/admin/scoring/nodes", "Bearer (Admin)", "CRUD nœuds", "app/api/admin/scoring/nodes/route.ts", "✅ Actif"),
        ("Admin Scoring", "POST/DELETE", "/api/admin/scoring/nodes/{id}/options", "Bearer (Admin)", "CRUD options nœud", "app/api/admin/scoring/nodes/[nodeId]/options/route.ts", "✅ Actif"),
        ("Admin Scoring", "POST/DELETE", "/api/admin/scoring/nodes/{id}/ranges", "Bearer (Admin)", "CRUD plages nœud", "app/api/admin/scoring/nodes/[nodeId]/ranges/route.ts", "✅ Actif"),
        ("Admin Scoring", "GET/POST", "/api/admin/scoring/overrides", "Bearer (Admin)", "Gestion surcharges", "app/api/admin/scoring/overrides/route.ts", "✅ Actif"),
        ("Admin Scoring", "GET/POST", "/api/admin/scoring/workflows", "Bearer (Admin)", "Gestion workflows", "app/api/admin/scoring/workflows/route.ts", "✅ Actif"),
        ("Admin Scoring", "GET/POST", "/api/admin/scoring/documents", "Bearer (Admin)", "Gestion documents", "app/api/admin/scoring/documents/route.ts", "✅ Actif"),
        ("Admin Scoring", "GET", "/api/admin/scoring/grid-stats", "Bearer (Admin)", "Statistiques grille", "app/api/admin/scoring/grid-stats/route.ts", "✅ Actif"),
        ("Admin Scoring", "POST", "/api/admin/scoring/validate-grid", "Bearer (Admin)", "Validation grille", "app/api/admin/scoring/validate-grid/route.ts", "✅ Actif"),
        # Admin Données
        ("Admin Données", "GET/POST", "/api/admin/scoring-criteria", "Bearer (Admin)", "CRUD critères", "app/api/admin/scoring-criteria/route.ts", "✅ Actif"),
        ("Admin Données", "GET/POST", "/api/admin/field-configurations", "Bearer (Admin)", "Config champs formulaires", "app/api/admin/field-configurations/route.ts", "✅ Actif"),
        ("Admin Données", "GET/POST", "/api/admin/countries", "Bearer (Admin)", "CRUD pays", "app/api/admin/countries/route.ts", "✅ Actif"),
        ("Admin Données", "GET", "/api/admin/config/country-risk-mode", "Bearer (Admin)", "Mode risque pays", "app/api/admin/config/country-risk-mode/route.ts", "✅ Actif"),
        ("Admin Données", "GET/POST/PUT/DELETE", "/api/admin/users", "Bearer (Admin)", "CRUD utilisateurs", "app/api/admin/users/route.ts", "✅ Actif"),
        ("Admin Données", "GET", "/api/admin/diagnostic/full", "Bearer (Admin)", "Diagnostic complet", "app/api/admin/diagnostic/full/route.ts", "✅ Actif"),
        # Utilitaires
        ("Utilitaires", "GET", "/api/health", "Non", "Health check", "app/api/health/route.ts", "✅ Actif"),
        ("Utilitaires", "GET", "/api/audit", "Bearer", "Logs audit", "app/api/audit/route.ts", "✅ Actif"),
        ("Utilitaires", "POST", "/api/alerts/create", "Bearer", "Créer alerte", "app/api/alerts/create/route.ts", "✅ Actif"),
        # Debug (à désactiver)
        ("⚠️ Debug", "GET/POST", "/api/debug/login", "NON ← DANGER", "Login sans auth (debug)", "app/api/debug/login/route.ts", "❌ À désactiver"),
        ("⚠️ Debug", "GET", "/api/debug/users", "NON ← DANGER", "Liste users sans auth", "app/api/debug/users/route.ts", "❌ À désactiver"),
        ("⚠️ Init", "POST", "/api/init-test-user", "NON ← DANGER", "Créer user test", "app/api/init-test-user/route.ts", "❌ À sécuriser"),
        ("⚠️ Init", "POST", "/api/init/set-admin-password", "Token secret", "Définir mdp admin", "app/api/init/set-admin-password/route.ts", "⚠️ Sécuriser"),
    ]

    METH_COLORS = {
        "GET": ("166534", "DCFCE7"),
        "POST": ("1E40AF", LIGHT_BLUE),
        "PUT": ("92400E", ORANGE_LIGHT),
        "DELETE": (RED, RED_LIGHT),
        "GET/POST": ("166534", "DCFCE7"),
        "GET/POST/PUT/DELETE": ("7C3AED", "EDE9FE"),
        "POST/DELETE": ("1E40AF", LIGHT_BLUE),
    }

    for row_idx, route in enumerate(routes, start=3):
        cat, meth, ep, auth, desc, fichier, statut = route
        bg_row = GRAY_LIGHT if row_idx % 2 == 0 else WHITE

        cat_bg = RED_LIGHT if "⚠️" in cat else bg_row
        c = ws.cell(row=row_idx, column=1, value=cat)
        style_cell(c, cat_bg, DARK_BLUE, True, 9)

        fg_m, bg_m = METH_COLORS.get(meth, (GRAY, GRAY_LIGHT))
        c = ws.cell(row=row_idx, column=2, value=meth)
        style_cell(c, bg_m, fg_m, True, 9, "center")

        c = ws.cell(row=row_idx, column=3, value=ep)
        style_cell(c, bg_row, MID_BLUE, False, 9, "left")

        auth_bg = RED_LIGHT if "DANGER" in auth else bg_row
        auth_fg = RED if "DANGER" in auth else "475569"
        c = ws.cell(row=row_idx, column=4, value=auth)
        style_cell(c, auth_bg, auth_fg, "DANGER" in auth, 9, "center")

        c = ws.cell(row=row_idx, column=5, value=desc)
        style_cell(c, bg_row, "1E293B", False, 9, "left")

        c = ws.cell(row=row_idx, column=6, value=fichier)
        style_cell(c, bg_row, GRAY, False, 8, "left")

        stat_fg = GREEN if "✅" in statut else (RED if "❌" in statut else ORANGE)
        stat_bg = GREEN_LIGHT if "✅" in statut else (RED_LIGHT if "❌" in statut else ORANGE_LIGHT)
        c = ws.cell(row=row_idx, column=7, value=statut)
        style_cell(c, stat_bg, stat_fg, True, 9, "center")

        ws.row_dimensions[row_idx].height = 18

    col_width(ws, {
        "A": 16, "B": 18, "C": 45, "D": 18,
        "E": 35, "F": 42, "G": 15
    })
    freeze(ws, "A3")
    ws.auto_filter.ref = f"A2:G{len(routes)+2}"


# ══════════════════════════════════════════════════════════════════
# 5. ARCHITECTURE SCORING
# ══════════════════════════════════════════════════════════════════
def sheet_scoring(wb):
    ws = wb.create_sheet("⚙️ Moteur de Scoring")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ARCHITECTURE MOTEUR DE SCORING — DOMAINES & CRITÈRES"
    style_header(c, DARK_BLUE, WHITE, True, 14)
    ws.row_dimensions[1].height = 35

    headers = ["Domaine", "Poids", "Sous-domaine / Critère", "Type Score",
               "Méthode Agrégation", "Conformité"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, MID_BLUE, WHITE, True, 10)

    domains = [
        ("D1 – Sponsor & Shareholders", "10%", "Sponsor Strength", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D1 – Sponsor & Shareholders", "10%", "Shareholder Structure", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D1 – Sponsor & Shareholders", "10%", "Governance Quality", "OPTION_SCORE", "WEIGHTED_AVERAGE", "IFC PS 1"),
        ("D2 – Project Characteristics", "10%", "Technical Complexity", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D2 – Project Characteristics", "10%", "Location Risk", "RANGE_SCORE", "WEIGHTED_AVERAGE", "BAM"),
        ("D2 – Project Characteristics", "10%", "Project Size & Structure", "NUMERIC_DIRECT", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D3 – Construction Risk", "15%", "EPC Completion Risk", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D3 – Construction Risk", "15%", "Interfaces Risk", "OPTION_SCORE", "WEIGHTED_AVERAGE", "EBRD"),
        ("D3 – Construction Risk", "15%", "Insurance Coverage", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D4 – Market Risk", "10%", "Demand Risk", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D4 – Market Risk", "10%", "Price Risk", "RANGE_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D4 – Market Risk", "10%", "Competition Level", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D5 – Operational Risk", "10%", "O&M Capability", "OPTION_SCORE", "WEIGHTED_AVERAGE", "EBRD"),
        ("D5 – Operational Risk", "10%", "Operational Stability", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D5 – Operational Risk", "10%", "Cost Control", "RANGE_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D6 – Counterparty Risk", "10%", "Offtaker Risk", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D6 – Counterparty Risk", "10%", "Supplier Risk", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D6 – Counterparty Risk", "10%", "Financial Partners", "OPTION_SCORE", "WEIGHTED_AVERAGE", "BAM"),
        ("D7 – Financial Structure & Cash Flow", "15%", "DSCR (Debt Service Coverage Ratio)", "RANGE_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D7 – Financial Structure & Cash Flow", "15%", "LLCR (Loan Life Coverage Ratio)", "RANGE_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D7 – Financial Structure & Cash Flow", "15%", "Equity / Leverage", "RANGE_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D7 – Financial Structure & Cash Flow", "15%", "Cash Flow Predictability", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D8 – Legal & Documentation", "10%", "Contractual Framework", "OPTION_SCORE", "WEIGHTED_AVERAGE", "BAM"),
        ("D8 – Legal & Documentation", "10%", "Security Package", "OPTION_SCORE", "WEIGHTED_AVERAGE", "Bâle III"),
        ("D8 – Legal & Documentation", "10%", "Legal Environment", "OPTION_SCORE", "WEIGHTED_AVERAGE", "BAM"),
        ("D9 – ESG & Climate Risk", "10%", "Environmental Impact (IFC PS 1-8)", "OPTION_SCORE", "WEIGHTED_AVERAGE", "IFC PS"),
        ("D9 – ESG & Climate Risk", "10%", "Social Impact", "OPTION_SCORE", "WEIGHTED_AVERAGE", "EBRD ESP"),
        ("D9 – ESG & Climate Risk", "10%", "Governance & Ethics", "OPTION_SCORE", "WEIGHTED_AVERAGE", "IFC PS 1"),
        ("D9 – ESG & Climate Risk", "10%", "Climate Risk (Physical & Transition)", "OPTION_SCORE", "WEIGHTED_AVERAGE", "TCFD"),
    ]

    DOMAIN_COLORS = {
        "D1": ("1E3A8A", LIGHT_BLUE),
        "D2": ("065F46", "D1FAE5"),
        "D3": ("7C2D12", "FFEDD5"),
        "D4": ("4C1D95", "EDE9FE"),
        "D5": ("0C4A6E", "E0F2FE"),
        "D6": ("422006", "FEF3C7"),
        "D7": ("14532D", "DCFCE7"),
        "D8": ("1E1B4B", "EEF2FF"),
        "D9": ("064E3B", "ECFDF5"),
    }

    for row_idx, domain in enumerate(domains, start=3):
        dom, poids, crit, score_type, agg, conf = domain
        key = dom[:2]
        fg_d, bg_d = DOMAIN_COLORS.get(key, (GRAY, GRAY_LIGHT))

        c = ws.cell(row=row_idx, column=1, value=dom)
        style_cell(c, bg_d, fg_d, True, 10)

        c = ws.cell(row=row_idx, column=2, value=poids)
        style_cell(c, bg_d, fg_d, True, 10, "center")

        c = ws.cell(row=row_idx, column=3, value=crit)
        style_cell(c, GRAY_LIGHT if row_idx%2==0 else WHITE, "1E293B", False, 10)

        c = ws.cell(row=row_idx, column=4, value=score_type)
        style_cell(c, LIGHT_BLUE, MID_BLUE, False, 9, "center")

        c = ws.cell(row=row_idx, column=5, value=agg)
        style_cell(c, GREEN_LIGHT, GREEN, False, 9, "center")

        c = ws.cell(row=row_idx, column=6, value=conf)
        style_cell(c, ORANGE_LIGHT, ORANGE, True, 9, "center")

        ws.row_dimensions[row_idx].height = 20

    col_width(ws, {"A": 36, "B": 8, "C": 38, "D": 18, "E": 20, "F": 16})
    freeze(ws, "A3")

    # Tableau grades
    ws.merge_cells("H1:M1")
    c = ws["H1"]
    c.value = "ÉCHELLE DE NOTATION AAA → D"
    style_header(c, DARK_BLUE, WHITE, True, 12)

    grade_headers = ["Grade", "Score Min", "Score Max", "PD Min", "PD Max", "Classe Risque"]
    for col, h in enumerate(grade_headers, 8):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, MID_BLUE, WHITE, True, 10)

    grades = [
        ("AAA", 95, 100, "0.01%", "0.05%", "Faible"),
        ("AA",  85, 94,  "0.05%", "0.10%", "Faible"),
        ("A",   75, 84,  "0.10%", "0.30%", "Modéré"),
        ("BBB+",70, 74,  "0.30%", "0.50%", "Modéré"),
        ("BBB", 65, 69,  "0.50%", "0.80%", "Modéré"),
        ("BBB-",60, 64,  "0.80%", "1.20%", "Élevé"),
        ("BB+", 55, 59,  "1.20%", "2.00%", "Élevé"),
        ("BB",  50, 54,  "2.00%", "3.50%", "Élevé"),
        ("B",   45, 49,  "5.00%", "8.00%", "Très Élevé"),
        ("CCC", 35, 44,  "10.0%", "15.0%", "Très Élevé"),
        ("CC",  25, 34,  "15.0%", "20.0%", "Très Élevé"),
        ("C",   15, 24,  "20.0%", "50.0%", "Très Élevé"),
        ("D",    0, 14,  "50.0%", "100%",  "Défaut"),
    ]
    GRADE_BG = {
        "AAA": "14532D", "AA": "166534", "A": "15803D",
        "BBB+": "1D4ED8", "BBB": "2563EB", "BBB-": "3B82F6",
        "BB+": "D97706", "BB": "F59E0B",
        "B": "EA580C",
        "CCC": "B91C1C", "CC": "DC2626", "C": "EF4444", "D": "7F1D1D"
    }

    for row_idx, grade in enumerate(grades, start=3):
        g, mn, mx, pd_mn, pd_mx, cls = grade
        bg_g = GRADE_BG.get(g, GRAY)

        c = ws.cell(row=row_idx, column=8, value=g)
        c.fill = hfill(bg_g); c.font = hfont(WHITE, True, 11)
        c.alignment = halign("center"); c.border = hborder()

        for col_off, val in enumerate([mn, mx, pd_mn, pd_mx, cls], 1):
            c = ws.cell(row=row_idx, column=8+col_off, value=val)
            style_cell(c, GRAY_LIGHT if row_idx%2==0 else WHITE, "1E293B", False, 10, "center")

        ws.row_dimensions[row_idx].height = 20

    col_width(ws, {"H": 8, "I": 10, "J": 10, "K": 10, "L": 10, "M": 14})


# ══════════════════════════════════════════════════════════════════
# 6. RECOMMANDATIONS
# ══════════════════════════════════════════════════════════════════
def sheet_reco(wb):
    ws = wb.create_sheet("💡 Recommandations")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "PLAN D'ACTION & RECOMMANDATIONS PRIORISÉES"
    style_header(c, DARK_BLUE, WHITE, True, 14)
    ws.row_dimensions[1].height = 35

    headers = ["#", "Priorité", "Catégorie", "Action recommandée",
               "Effort", "Impact", "Délai suggéré"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, MID_BLUE, WHITE, True, 10)

    recos = [
        (1, "🔴 P0 – Urgent", "Sécurité",
         "Désactiver routes /api/debug/* et /api/init-test-user en production (check NODE_ENV)",
         "Faible (1h)", "Critique", "Immédiat"),
        (2, "🔴 P0 – Urgent", "Sécurité",
         "Configurer DATABASE_URL, JWT_SECRET et variables Supabase sur Vercel",
         "Faible (30min)", "Critique", "Immédiat"),
        (3, "🔴 P0 – Urgent", "Sécurité",
         "Fail-fast si JWT_SECRET absent en production (throw Error au démarrage)",
         "Faible (30min)", "Critique", "Immédiat"),
        (4, "🟠 P1 – Haute", "Sécurité",
         "Ajouter rate limiting sur /api/auth/login (max 5 tentatives/minute par IP)",
         "Moyen (2-4h)", "Haute", "Sprint 1"),
        (5, "🟠 P1 – Haute", "Fonctionnel",
         "Compléter l'évaluateur de formules FORMULA_SCORE avec expr-eval ou mathjs",
         "Moyen (4-8h)", "Haute", "Sprint 1"),
        (6, "🟠 P1 – Haute", "Cohérence",
         "Standardiser les réponses API avec lib/api-response.ts sur tous les endpoints",
         "Moyen (4-8h)", "Haute", "Sprint 1"),
        (7, "🟠 P1 – Haute", "Fonctionnel",
         "Compléter UI export Excel (/app/exports) avec sélection colonnes et download",
         "Moyen (4-8h)", "Haute", "Sprint 1"),
        (8, "🟡 P2 – Moyenne", "Sécurité",
         "Valider les champs JSON (conditionExpression, uiSchemaJson) avec schémas Zod",
         "Moyen (4-8h)", "Moyenne", "Sprint 2"),
        (9, "🟡 P2 – Moyenne", "Auth",
         "Configurer et tester le flux OAuth complet (Google/GitHub provider)",
         "Élevé (1-2j)", "Moyenne", "Sprint 2"),
        (10, "🟡 P2 – Moyenne", "Performance",
         "Activer React Query pour le cache des appels API répétitifs (scoring, projects)",
         "Moyen (4-8h)", "Moyenne", "Sprint 2"),
        (11, "🟡 P2 – Moyenne", "Fonctionnel",
         "Compléter les transformations Data Binding (FORMAT, MAP_VALUE, AGGREGATE)",
         "Élevé (2-3j)", "Moyenne", "Sprint 2"),
        (12, "🟢 P3 – Faible", "Qualité",
         "Ajouter tests unitaires sur scoring-engine et rule-engine (Jest/Vitest)",
         "Élevé (3-5j)", "Faible", "Sprint 3"),
        (13, "🟢 P3 – Faible", "Compliance",
         "Auditer conformité GDPR: vérifier chiffrement données sensibles au repos",
         "Élevé (2-3j)", "Moyenne", "Sprint 3"),
        (14, "🟢 P3 – Faible", "DevOps",
         "Ajouter monitoring Sentry pour tracking erreurs runtime en production",
         "Moyen (2-4h)", "Faible", "Sprint 3"),
        (15, "🟢 P3 – Faible", "Performance",
         "Implémenter pagination côté serveur sur toutes les listes (projets, évaluations)",
         "Moyen (4-8h)", "Faible", "Sprint 4"),
    ]

    PRIO_STYLES = {
        "🔴 P0 – Urgent": (RED, RED_LIGHT),
        "🟠 P1 – Haute":  (ORANGE, ORANGE_LIGHT),
        "🟡 P2 – Moyenne":("7C6F00", "FEFCE8"),
        "🟢 P3 – Faible": (GREEN, GREEN_LIGHT),
    }

    for row_idx, reco in enumerate(recos, start=3):
        num, prio, cat, action, effort, impact, delai = reco
        bg_row = GRAY_LIGHT if row_idx % 2 == 0 else WHITE
        fg_p, bg_p = PRIO_STYLES.get(prio, (GRAY, GRAY_LIGHT))

        ws.cell(row=row_idx, column=1, value=num).fill = hfill(bg_row)
        ws.cell(row=row_idx, column=1).font = hfont(GRAY, False, 9)
        ws.cell(row=row_idx, column=1).alignment = halign("center")
        ws.cell(row=row_idx, column=1).border = hborder()

        c = ws.cell(row=row_idx, column=2, value=prio)
        style_cell(c, bg_p, fg_p, True, 9, "center")

        c = ws.cell(row=row_idx, column=3, value=cat)
        style_cell(c, bg_row, DARK_BLUE, True, 9)

        c = ws.cell(row=row_idx, column=4, value=action)
        style_cell(c, bg_row, "1E293B", False, 10, "left", True)

        c = ws.cell(row=row_idx, column=5, value=effort)
        style_cell(c, bg_row, GRAY, False, 9, "center")

        imp_bg = GREEN_LIGHT if impact == "Critique" else (ORANGE_LIGHT if impact == "Haute" else bg_row)
        imp_fg = GREEN if impact == "Critique" else (ORANGE if impact == "Haute" else "475569")
        c = ws.cell(row=row_idx, column=6, value=impact)
        style_cell(c, imp_bg, imp_fg, True, 9, "center")

        c = ws.cell(row=row_idx, column=7, value=delai)
        style_cell(c, bg_row, GRAY, False, 9, "center")

        ws.row_dimensions[row_idx].height = 35

    col_width(ws, {"A": 4, "B": 17, "C": 14, "D": 65, "E": 16, "F": 11, "G": 14})
    freeze(ws, "A3")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    # Supprimer feuille par défaut
    wb.remove(wb.active)

    sheet_resume(wb)
    sheet_specs(wb)
    sheet_scoring(wb)
    sheet_bugs(wb)
    sheet_api(wb)
    sheet_reco(wb)

    output = "/home/user/pf-scoring-v7claude/RAPPORT_AUDIT_PF_SCORING_V7.xlsx"
    wb.save(output)
    print(f"✅ Rapport généré: {output}")
    print(f"   Onglets: {[ws.title for ws in wb.worksheets]}")

if __name__ == "__main__":
    main()
